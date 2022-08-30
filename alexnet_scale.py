from __future__ import division
import torch
import torch.nn as nn
import torch.optim as optim
from torch.optim.lr_scheduler import LambdaLR
import numpy as np
import torchvision
from torchvision import datasets, models, transforms
import matplotlib.pyplot as plt
import time
import os
import copy
import json
import shutil
from PIL import Image
from tqdm import tqdm
import load_flare
import argparse
from openpyxl import load_workbook
import pandas as pd
from torch.utils.data import DataLoader,WeightedRandomSampler
from itertools import cycle
from sklearn.metrics import roc_curve, auc, f1_score, precision_recall_curve, average_precision_score
from scipy import interp

parser = argparse.ArgumentParser(description='Process some integers.')
print("PyTorch Version: ",torch.__version__)
print("Torchvision Version: ",torchvision.__version__)


# Models to choose from [resnet, alexnet, vgg, squeezenet, densenet, inception]
parser.add_argument('--model_name', type=str, default='res18')
parser.add_argument('--num_classes', type=int, default=2)
parser.add_argument('--batch_size', type=int, default=256)
parser.add_argument('--num_epochs', type=int, default=40)
parser.add_argument('--resize_scale', type=int, default=512)
parser.add_argument('--reduce_resolution', type=int, default=1)
parser.add_argument('--data_aug', type=bool, default=True)
parser.add_argument('--save_dir', type=str, default='./rst/test/')
parser.add_argument('--cross', type=bool, default=False)
parser.add_argument('--cross_test_year', type=int, default=2010)
# Flag for feature extracting. When False, we finetune the whole model,
# when True we only update the reshaped layer params
parser.add_argument('--feature_extract', type=bool, default=False)
args = parser.parse_args()
wb = pd.DataFrame()
wb.to_excel("alex_result_"+str(args.reduce_resolution)+".xlsx")
def train_model(model, dataloaders, criterion, optimizer, scheduler, num_epochs, is_inception=False,rst_dir='./rst/exp'):
    val_acc_history = []
    lr_init = optimizer.param_groups[0]['lr']
    for epoch in range(num_epochs):

        print('Epoch {}/{}'.format(epoch, num_epochs - 1))
        print('-' * 10)
        adjust_learning_rate(optimizer, epoch, lr_init)

        # Each epoch has a training and validation phase
        TP = 0
        TN = 0
        FN = 0
        FP = 0
        run_time=time.time()

        f_log = open(os.path.join(rst_dir, 'logs'), 'a')


        model.train()  # Set model to training mode
        print('lr = %f'%(optimizer.param_groups[0]['lr']))

        running_loss = 0.0
        running_corrects = 0

        for inputs, gt,inputs_neg,gt_neg in dataloaders['train']:
            inputs = torch.cat((inputs, inputs_neg), 0)
            gt = torch.cat((gt, gt_neg), 0)
            inputs = inputs.cuda()
            gt = np.array(gt)
            labels = torch.from_numpy(gt)
            labels = labels.cuda()
            optimizer.zero_grad()


            outputs = model(inputs)
            loss = criterion(outputs, labels)

            _, preds = torch.max(outputs, 1)

            loss.backward()
            optimizer.step()

            running_loss += loss.item() * inputs.size(0)
            running_corrects += torch.sum(preds == labels.data)
            a = running_corrects.double()

            TP += ((preds == 1) & (labels.data == 1)).cpu().sum().to(a)
            TN += ((preds == 0) & (labels.data == 0)).cpu().sum().to(a)
            FN += ((preds == 0) & (labels.data == 1)).cpu().sum().to(a)
            FP += ((preds == 1) & (labels.data == 0)).cpu().sum().to(a)

        epoch_loss = running_loss / (len(dataloaders['train'].dataset)*2)
        epoch_acc = running_corrects.double() / (len(dataloaders['train'].dataset)*2)

        TPR = TP /(TP + FN)
        TNR = TN /(TN + FP)
        FPR = FP /(FP + TN)
        FNR = FN /(TP + FN)
        acc = (TP + TN) / (TP + TN + FP + FN)

        if scheduler!=None:
            scheduler.step()
        torch.save(model, os.path.join(rst_dir, str(epoch) + '_' + str(float('%.2f' % (epoch_acc*100))) + '.pth'))

        run_time=time.time()-run_time
        print('{} Loss: {:.4f} Acc: {:.4f} Time: {:.2f}min'.format('train', epoch_loss, epoch_acc,run_time/60.0))
        print('TP: {:4f}, TN: {:4f}, FN: {:4f}, FP: {:4f}'.format(TP, TN, FN, FP))
        print('TPR: {:4f}, TNR: {:4f}, FPR: {:4f}, FNR: {:4f}, acc: {:4f}'.format(TPR, TNR, FPR, FNR, acc))
        print("\n")
        log1 = 'TP: {:4f}, TN: {:4f}, FN: {:4f}, FP: {:4f}'.format(TP, TN, FN, FP)
        log = '{} Loss: {:.4f} Acc: {:.4f} TPR: {:4f}, TNR: {:4f}, FPR: {:4f}, FNR: {:4f}'.format('train', epoch_loss, epoch_acc,TPR, TNR, FPR, FNR)
        #print(log)
        wb = load_workbook("alex_result_" + str(args.reduce_resolution) + ".xlsx")
        wb1 = wb.active
        wb1.cell(epoch+1, 1, epoch_loss)
        wb1.cell(epoch+1, 2, float(epoch_acc.cpu().detach().numpy()))
        wb1.cell(epoch+1, 3, float(TPR.cpu().detach().numpy()))
        wb1.cell(epoch+1, 4, float(TNR.cpu().detach().numpy()))
        wb.save("alex_result_" + str(args.reduce_resolution) + ".xlsx")
        f_log.writelines(str(epoch) + '_train:' '\n')
        f_log.writelines(log1 + '\n')
        f_log.writelines(str(epoch) + ': ' + log + '\n')
        f_log.close()

        test_model(model, dataloaders, criterion, optimizer, rst_dir, epoch)

    return model, val_acc_history


def test_model(model, dataloaders, criterion, optimizer,rst_dir, num_epochs):
    epoch=num_epochs
    TP = 0
    TN = 0
    FN = 0
    FP = 0
    run_time = time.time()

    f_log = open(os.path.join(rst_dir, 'logs'), 'a')
    model.eval()  # Set model to evaluate mode

    running_loss = 0.0
    running_corrects = 0

    y_score = []
    y_test = []
    for inputs, gt in dataloaders['val']:
        inputs = inputs.cuda()
        gt = np.array(gt)
        labels = torch.from_numpy(gt)
        labels = labels.cuda()
        optimizer.zero_grad()


        outputs = model(inputs)
        out = torch.nn.functional.softmax(outputs, dim=1)
        out = np.array(out.cpu().detach().numpy())
        out_right = out[:,1]

        loss = criterion(outputs, labels)

        _, preds = torch.max(outputs, 1)

        # statistics
        running_loss += loss.item() * inputs.size(0)
        running_corrects += torch.sum(preds == labels.data)
        a = running_corrects.double()

        TP += ((preds == 1) & (labels.data == 1)).cpu().sum().to(a)
        TN += ((preds == 0) & (labels.data == 0)).cpu().sum().to(a)
        FN += ((preds == 0) & (labels.data == 1)).cpu().sum().to(a)
        FP += ((preds == 1) & (labels.data == 0)).cpu().sum().to(a)

        epoch_loss = running_loss / len(dataloaders['val'].dataset)
        epoch_acc = running_corrects.double() / len(dataloaders['val'].dataset)

        TPR = TP / (TP + FN)
        TNR = TN / (TN + FP)
        FPR = FP / (FP + TN)
        FNR = FN / (TP + FN)
        acc = (TP + TN) / (TP + TN + FP + FN)

        y_score.extend(out_right)
        y_test.extend(labels.cpu().numpy())

        run_time = time.time() - run_time

    print('{} Loss: {:.4f} Acc: {:.4f} Time: {:.2f}min'.format('val', epoch_loss, epoch_acc, run_time / 60.0))
    print('TP: {:4f}, TN: {:4f}, FN: {:4f}, FP: {:4f}'.format(TP, TN, FN, FP))
    print('TPR: {:4f}, TNR: {:4f}, FPR: {:4f}, FNR: {:4f}, acc: {:4f}'.format(TPR, TNR, FPR, FNR, acc))
    print("\n")
    log1 = 'TP: {:4f}, TN: {:4f}, FN: {:4f}, FP: {:4f}'.format(TP, TN, FN, FP)
    log = '{} Loss: {:.4f} Acc: {:.4f} TPR: {:4f}, TNR: {:4f}, FPR: {:4f}, FNR: {:4f}'.format('train', epoch_loss,
                                                                                              epoch_acc, TPR, TNR, FPR,
                                                                                              FNR)
    # print(log)
    wb = load_workbook("alex_result_" + str(args.reduce_resolution) + ".xlsx")
    wb1 = wb.active
    wb1.cell(epoch + 1, 5, epoch_loss)
    wb1.cell(epoch + 1, 6, float(epoch_acc.cpu().detach().numpy()))
    wb1.cell(epoch + 1, 7, float(TPR.cpu().detach().numpy()))
    wb1.cell(epoch + 1, 8, float(TNR.cpu().detach().numpy()))
    wb.save("alex_result_" + str(args.reduce_resolution) + ".xlsx")
    f_log.writelines(str(epoch) + '_val:' '\n')
    f_log.writelines(log1 + '\n')
    f_log.writelines(str(epoch) + ': ' + log + '\n')
    f_log.close()
    np.save(rst_dir + str(epoch) + '_df', [y_test, y_score])

def set_parameter_requires_grad(model, feature_extracting):
    if feature_extracting:
        for param in model.parameters():
            param.requires_grad = False

def adjust_learning_rate(optimizer, epoch, lr):
    """Sets the learning rate to the initial LR decayed by 10 every 2 epochs"""
    lr *= (0.1 ** (epoch // 15))
    for param_group in optimizer.param_groups:
        param_group['lr'] = lr
if __name__ == '__main__':
    """ Alexnet
            """
    preTrain=True


    #rst_dir='./rst/'+'alex_exp'+str(os.listdir('./rst').__len__())
    rst_dir = args.save_dir
    os.mkdir(rst_dir)
    shutil.copy(__file__, os.path.join(rst_dir,os.path.split(__file__)[1]))
    print(rst_dir)
    print(os.environ.get('CUDA_VISIBLE_DEVICES'))

    model_ft = models.alexnet(pretrained=preTrain)
    #set_parameter_requires_grad(model_ft, args.feature_extract)
    num_ftrs = model_ft.classifier[6].in_features
    model_ft.classifier[6] = nn.Linear(num_ftrs, args.num_classes)


    if args.data_aug == False:
        data_transforms = {
            'train': transforms.Compose([
                #transforms.ToPILImage(),
                transforms.Resize((args.resize_scale,args.resize_scale),interpolation=Image.BICUBIC),
                transforms.RandomHorizontalFlip(),
                transforms.ToTensor(),
                transforms.Normalize((0.485, 0.456, 0.406), (0.229, 0.224, 0.225))
            ]),
            'val': transforms.Compose([
                #transforms.ToPILImage(),
                transforms.Resize((args.resize_scale,args.resize_scale),interpolation=Image.BICUBIC),
                transforms.ToTensor(),
                transforms.Normalize((0.485, 0.456, 0.406), (0.229, 0.224, 0.225))
            ]),
        }
    else:
        data_transforms = {
            'train': transforms.Compose([
                #transforms.ToPILImage(),
                #transforms.Scale(args.reduce_resolution, interpolation=2),
                #transforms.Scale(args.resize_scale, interpolation=2),
                #transforms.CenterCrop((args.resize_scale, args.resize_scale)),
                #transforms.Resize((args.reduce_resolution, args.reduce_resolution), interpolation=Image.BICUBIC),
                transforms.Resize((args.resize_scale, args.resize_scale),interpolation=Image.BICUBIC),
                transforms.RandomHorizontalFlip(),
                transforms.RandomVerticalFlip(p=0.5),
                transforms.RandomRotation([90, 270]),
                transforms.ToTensor(),
                transforms.Normalize((0.485, 0.456, 0.406), (0.229, 0.224, 0.225))
            ]),
            'val': transforms.Compose([
                #transforms.ToPILImage(),
                #transforms.Resize((args.reduce_resolution, args.reduce_resolution), interpolation=Image.BICUBIC),
                transforms.Resize((args.resize_scale, args.resize_scale), interpolation=Image.BICUBIC),
                transforms.ToTensor(),
                transforms.Normalize((0.485, 0.456, 0.406), (0.229, 0.224, 0.225))
            ]),
        }

    print("Initializing Datasets and Dataloaders...")

    # Create training and validation datasets

    # image_datasets = {x: datasets.ImageFolder(os.path.join(data_dir, x), data_transforms[x]) for x in ['train', 'val']}
    image_datasets = {'train':load_flare.flare(args.reduce_resolution, args.cross, args.cross_test_year, split='train',transform=data_transforms['train']),
                      'val':load_flare.flare_test(args.reduce_resolution,args.cross, args.cross_test_year,split='test',transform=data_transforms['val'])}
    # Create training and validation dataloaders

    dataloaders_dict = {
        x: torch.utils.data.DataLoader(image_datasets[x], batch_size=args.batch_size, shuffle=True, num_workers=0,pin_memory=True) for x in
        ['train', 'val']}

    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    model_ft = model_ft.to(device)


    #params_to_update = model_ft.parameters()
    print("Params to learn:")
    nofreeze_layers = ( "features.8", "features.10", "classifier.1")
    for name, param in model_ft.named_parameters():
        param.requires_grad = True
        print("\t", name)

    # Observe that all parameters are being optimized
    optimizer_type='sgd'
    if optimizer_type=='sgd':
        optimizer_ft = optim.SGD(filter(lambda p: p.requires_grad, model_ft.parameters()), lr=0.0001, momentum=0.9)
        #StepLR = torch.optim.lr_scheduler.StepLR(optimizer_ft, step_size=400, gamma=0.4)
    elif optimizer_type=='adam':
        optimizer_ft = optim.AdamW(filter(lambda p: p.requires_grad, model_ft.parameters()), lr=0.0001, betas=(0.9, 0.999), eps=1e-08, weight_decay=0)
    #scheduler = LambdaLR(optimizer_ft, lr_lambda=lambda epoch: lr[epoch]/1.0)
    scheduler = None

    # Setup the loss fxn
    criterion = nn.CrossEntropyLoss()

    infos={
        'preTrain':preTrain,
        'num_epochs':args.num_epochs,
        'batch_size':args.batch_size,
        'optim':optimizer_type,
        'lr':lr if scheduler!=None else 'None',
    }
    print(infos)
    json.dump(infos,open(os.path.join(rst_dir,'infos.txt'),'w'))

    # Train and evaluate
    model_ft, hist = train_model(model_ft, dataloaders_dict, criterion, optimizer_ft, scheduler, num_epochs=args.num_epochs,
                                 is_inception=(args.model_name == "inception"),rst_dir=rst_dir)
